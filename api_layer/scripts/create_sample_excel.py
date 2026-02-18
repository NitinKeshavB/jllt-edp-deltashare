#!/usr/bin/env python3
"""
Generate full sample Excel share pack file matching the YAML template and parser schema.

Usage:
    cd api_layer && python scripts/create_sample_excel.py

Output:
    sharepack_templates/sample_sharepack.xlsx
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

# Output under sharepack_templates (same folder as sample_sharepack.yaml)
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "sharepack_templates"
OUTPUT_FILE = OUTPUT_DIR / "sample_sharepack.xlsx"


def create_sample_excel():
    """Create full sample Excel sharepack matching YAML template and parser columns."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # -------------------------------------------------------------------------
    # Sheet 1: Metadata (key-value; parser uses min_row=2, A=key B=value)
    # -------------------------------------------------------------------------
    ws_meta = wb.create_sheet("Metadata", 0)
    ws_meta["A1"] = "Field"
    ws_meta["B1"] = "Value"
    ws_meta["A1"].fill = header_fill
    ws_meta["B1"].fill = header_fill
    ws_meta["A1"].font = header_font
    ws_meta["B1"].font = header_font

    metadata_rows = [
        ("requestor", "test.user@jll.com"),
        ("project_name", "SharePack Demo - Q1 2024"),
        ("business_line", "Data Platform Engineering"),
        ("strategy", "NEW"),
        ("description", "Demo SharePack - full template"),
        ("delta_share_region", "AM"),
        ("configurator", "data-platform-team@jll.com"),
        ("approver", "analytics-leadership@jll.com"),
        ("executive_team", "data-governance-team@jll.com"),
        ("approver_status", "approved"),
        ("workspace_url", "https://adb-1234567890123456.12.azuredatabricks.net"),
        ("servicenow", "INC0012345"),
        ("version", "1.0"),
        ("contact_email", "test.user@jll.com"),
    ]
    for i, (k, v) in enumerate(metadata_rows, start=2):
        ws_meta[f"A{i}"] = k
        ws_meta[f"B{i}"] = v
    ws_meta.column_dimensions["A"].width = 22
    ws_meta.column_dimensions["B"].width = 55

    # -------------------------------------------------------------------------
    # Sheet 2: Recipients (parser: name, type, recipient, recipient_databricks_org, recipient_ips, token_expiry, token_rotation)
    # -------------------------------------------------------------------------
    ws_rec = wb.create_sheet("Recipients", 1)
    rec_headers = [
        "name",
        "type",
        "recipient",
        "recipient_databricks_org",
        "recipient_ips",
        "token_expiry",
        "token_rotation",
    ]
    for c, h in enumerate(rec_headers, start=1):
        cell = ws_rec.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font

    rec_data = [
        (
            "external_partner_recipient",
            "D2O",
            "partner@external-company.com",
            "",
            "203.0.113.0/24,198.51.100.50",
            90,
            False,
        ),
        (
            "internal_analytics_team",
            "D2D",
            "analytics-team@jll.com",
            "aws:us-west-2:a1b2c3d4-e5f6-7890-abcd-ef1234567890",
            "",
            30,
            False,
        ),
    ]
    for r, row in enumerate(rec_data, start=2):
        for c, val in enumerate(row, start=1):
            ws_rec.cell(row=r, column=c, value=val)
    for c in range(1, len(rec_headers) + 1):
        ws_rec.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 18

    # -------------------------------------------------------------------------
    # Sheet 3: Shares (parser: share_name, asset, recipient, ext_catalog_name, ext_schema_name, prefix_assetname, share_tags)
    # One row per (share, asset) or (share, asset, recipient); assets/recipients collected per share.
    # -------------------------------------------------------------------------
    ws_share = wb.create_sheet("Shares", 2)
    share_headers = [
        "share_name",
        "asset",
        "recipient",
        "ext_catalog_name",
        "ext_schema_name",
        "prefix_assetname",
        "share_tags",
    ]
    for c, h in enumerate(share_headers, start=1):
        cell = ws_share.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font

    share_data = [
        (
            "sales_analytics_share",
            "main_catalog.sales_schema.daily_sales",
            "external_partner_recipient",
            "analytics_prod",
            "shared_sales",
            "",
            "production,sales_analytics",
        ),
        (
            "sales_analytics_share",
            "main_catalog.sales_schema.customer_orders",
            "external_partner_recipient",
            "analytics_prod",
            "shared_sales",
            "",
            "production,sales_analytics",
        ),
        (
            "sales_analytics_share",
            "main_catalog.sales_schema.revenue_summary",
            "external_partner_recipient",
            "analytics_prod",
            "shared_sales",
            "",
            "production,sales_analytics",
        ),
        (
            "sales_analytics_share",
            "main_catalog.sales_schema.daily_sales",
            "internal_analytics_team",
            "analytics_prod",
            "shared_sales",
            "",
            "production,sales_analytics",
        ),
        (
            "operations_realtime_share",
            "operations_catalog.metrics_schema.system_health",
            "internal_analytics_team",
            "ops_analytics",
            "realtime_metrics",
            "",
            "production,operations,monitoring",
        ),
        (
            "operations_realtime_share",
            "operations_catalog.metrics_schema.error_logs",
            "internal_analytics_team",
            "ops_analytics",
            "realtime_metrics",
            "",
            "production,operations,monitoring",
        ),
        (
            "operations_realtime_share",
            "operations_catalog.metrics_schema.user_activity",
            "internal_analytics_team",
            "ops_analytics",
            "realtime_metrics",
            "",
            "production,operations,monitoring",
        ),
        (
            "compliance_audit_share",
            "compliance_catalog.audit_schema.transaction_logs",
            "external_partner_recipient",
            "compliance_prod",
            "audit_trail",
            "",
            "production,compliance,gdpr,sox",
        ),
        (
            "compliance_audit_share",
            "compliance_catalog.audit_schema.access_records",
            "external_partner_recipient",
            "compliance_prod",
            "audit_trail",
            "",
            "production,compliance,gdpr,sox",
        ),
        (
            "compliance_audit_share",
            "compliance_catalog.audit_schema.policy_changes",
            "external_partner_recipient",
            "compliance_prod",
            "audit_trail",
            "",
            "production,compliance,gdpr,sox",
        ),
    ]
    for r, row in enumerate(share_data, start=2):
        for c, val in enumerate(row, start=1):
            ws_share.cell(row=r, column=c, value=val)
    ws_share.column_dimensions["A"].width = 28
    ws_share.column_dimensions["B"].width = 50
    ws_share.column_dimensions["C"].width = 28
    ws_share.column_dimensions["D"].width = 18
    ws_share.column_dimensions["E"].width = 18
    ws_share.column_dimensions["F"].width = 12
    ws_share.column_dimensions["G"].width = 35

    # -------------------------------------------------------------------------
    # Sheet 4: Pipelines (parser: share_name, name_prefix, source_asset, asset_name, schedule_type, cron, timezone, notification, tags, serverless, scd_type, key_columns)
    # -------------------------------------------------------------------------
    ws_pipe = wb.create_sheet("Pipelines", 3)
    pipe_headers = [
        "share_name",
        "name_prefix",
        "source_asset",
        "asset_name",
        "schedule_type",
        "cron",
        "timezone",
        "notification",
        "tags",
        "serverless",
        "scd_type",
        "key_columns",
    ]
    for c, h in enumerate(pipe_headers, start=1):
        cell = ws_pipe.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font

    pipeline_data = [
        (
            "sales_analytics_share",
            "sales_daily_sync",
            "main_catalog.sales_schema.daily_sales",
            "daily_sales",
            "CRON",
            "0 0 2 * * ?",
            "America/New_York",
            "analytics-team@jll.com,sales-ops@jll.com",
            "environment:production,dataset:daily_sales,owner:sales_team",
            False,
            "2",
            "sale_id,sale_date",
        ),
        (
            "sales_analytics_share",
            "sales_orders_sync",
            "main_catalog.sales_schema.customer_orders",
            "customer_orders",
            "CRON",
            "0 0 */6 * * ?",
            "UTC",
            "analytics-team@jll.com",
            "environment:production,dataset:customer_orders,owner:sales_team",
            True,
            "2",
            "order_id,customer_id,created_at",
        ),
        (
            "sales_analytics_share",
            "revenue_summary_sync",
            "main_catalog.sales_schema.revenue_summary",
            "revenue_summary",
            "CRON",
            "0 0 * * * ?",
            "America/New_York",
            "finance-team@jll.com,analytics-team@jll.com",
            "environment:production,dataset:revenue,owner:finance_team",
            True,
            "1",
            "",
        ),
        (
            "operations_realtime_share",
            "ops_health_stream",
            "operations_catalog.metrics_schema.system_health",
            "system_health",
            "CONTINUOUS",
            "",
            "UTC",
            "ops-team@jll.com",
            "environment:production,table:system_health,streaming:true",
            True,
            "1",
            "",
        ),
        (
            "operations_realtime_share",
            "ops_errors_sync",
            "operations_catalog.metrics_schema.error_logs",
            "error_logs",
            "CRON",
            "0 */15 * * * ?",
            "UTC",
            "ops-team@jll.com,sre-team@jll.com",
            "environment:production,table:error_logs",
            True,
            "1",
            "",
        ),
        (
            "operations_realtime_share",
            "ops_activity_sync",
            "operations_catalog.metrics_schema.user_activity",
            "user_activity",
            "CRON",
            "0 */30 * * * ?",
            "America/Los_Angeles",
            "analytics-team@jll.com",
            "environment:production,table:user_activity",
            True,
            "2",
            "user_id,activity_timestamp",
        ),
        (
            "compliance_audit_share",
            "compliance_txn_emea",
            "compliance_catalog.audit_schema.transaction_logs",
            "transaction_logs",
            "CRON",
            "0 0 1 * * ?",
            "Europe/London",
            "compliance-team@jll.com",
            "environment:production,region:EMEA,compliance:SOX",
            False,
            "2",
            "transaction_id,timestamp",
        ),
        (
            "compliance_audit_share",
            "compliance_access_apac",
            "compliance_catalog.audit_schema.access_records",
            "access_records",
            "CRON",
            "0 0 2 * * ?",
            "Asia/Tokyo",
            "compliance-team@jll.com",
            "environment:production,region:APAC,compliance:GDPR",
            True,
            "2",
            "access_id,user_id,timestamp",
        ),
        (
            "compliance_audit_share",
            "compliance_policy_stream",
            "compliance_catalog.audit_schema.policy_changes",
            "policy_changes",
            "CONTINUOUS",
            "",
            "UTC",
            "legal-team@jll.com,compliance-team@jll.com",
            "environment:production,compliance:ALL,priority:critical",
            True,
            "1",
            "",
        ),
    ]
    for r, row in enumerate(pipeline_data, start=2):
        for c, val in enumerate(row, start=1):
            ws_pipe.cell(row=r, column=c, value=val)
    ws_pipe.column_dimensions["A"].width = 28
    ws_pipe.column_dimensions["B"].width = 24
    ws_pipe.column_dimensions["C"].width = 52
    ws_pipe.column_dimensions["D"].width = 20
    ws_pipe.column_dimensions["E"].width = 12
    ws_pipe.column_dimensions["F"].width = 18
    ws_pipe.column_dimensions["G"].width = 20
    ws_pipe.column_dimensions["H"].width = 45
    ws_pipe.column_dimensions["I"].width = 50
    ws_pipe.column_dimensions["J"].width = 10
    ws_pipe.column_dimensions["K"].width = 8
    ws_pipe.column_dimensions["L"].width = 30

    # -------------------------------------------------------------------------
    # Save
    # -------------------------------------------------------------------------
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)

    print(f"Created: {OUTPUT_FILE}")
    print(f"Sheets: {', '.join(wb.sheetnames)}")
    print(f"Metadata rows: {len(metadata_rows)}")
    print(f"Recipients: {len(rec_data)}")
    print(f"Share rows: {len(share_data)}")
    print(f"Pipelines: {len(pipeline_data)}")


if __name__ == "__main__":
    create_sample_excel()
