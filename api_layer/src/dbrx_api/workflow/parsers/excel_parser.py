"""
Excel Parser

Parses Excel files into SharePackConfig models.
Expected sheets: Metadata, Recipients, Shares, Pipelines
"""

from typing import Union, BinaryIO
from io import BytesIO
from loguru import logger

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed - Excel parsing will not work")

from dbrx_api.workflow.models.share_pack import SharePackConfig


def parse_excel(file_content: Union[bytes, BinaryIO]) -> SharePackConfig:
    """
    Parse Excel file into SharePackConfig.

    Expected sheets:
    - Metadata: key-value pairs (A=key, B=value)
    - Recipients: columns = name, type, recipient, recipient_databricks_org, recipient_ips, token_expiry, token_rotation
    - Shares: columns = share_name, asset, recipient, ext_catalog_name, ext_schema_name, prefix_assetname, share_tags
    - Pipelines: columns = share_name, name_prefix, asset_name, schedule_type, cron, timezone, notification, tags, serverless, scd_type, key_columns

    Args:
        file_content: Excel file content as bytes or file-like object

    Returns:
        SharePackConfig instance

    Raises:
        ValueError: If required sheets are missing or data is invalid
        pydantic.ValidationError: If data doesn't match SharePackConfig schema
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl not installed - cannot parse Excel files")

    # Load workbook
    if isinstance(file_content, bytes):
        wb = openpyxl.load_workbook(BytesIO(file_content))
    else:
        wb = openpyxl.load_workbook(file_content)

    # Parse Metadata sheet
    metadata_dict = _parse_metadata_sheet(wb)

    # Parse Recipients sheet
    recipients = _parse_recipients_sheet(wb)

    # Parse Shares sheet
    shares = _parse_shares_sheet(wb)

    # Parse Pipelines sheet (optional)
    pipelines_by_share = _parse_pipelines_sheet(wb)

    # Attach pipelines to shares
    for share in shares:
        share_name = share["name"]
        if share_name in pipelines_by_share:
            share["pipelines"] = pipelines_by_share[share_name]
        else:
            share["pipelines"] = []

    # Construct final dict
    config_dict = {
        "metadata": metadata_dict,
        "recipient": recipients,
        "share": shares,
    }

    # Validate and convert to SharePackConfig
    try:
        config = SharePackConfig(**config_dict)
        logger.debug(
            f"Successfully parsed Excel: {len(config.recipient)} recipients, {len(config.share)} shares"
        )
        return config
    except Exception as e:
        logger.error(f"SharePackConfig validation error: {e}")
        raise


def _parse_metadata_sheet(wb) -> dict:
    """Parse Metadata sheet (key-value pairs)."""
    if "Metadata" not in wb.sheetnames:
        raise ValueError("Missing required sheet: Metadata")

    sheet = wb["Metadata"]
    metadata = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
        if row[0] is None:  # Skip empty rows
            continue
        key = str(row[0]).strip()
        value = row[1] if row[1] is not None else ""
        metadata[key] = value

    return metadata


def _parse_recipients_sheet(wb) -> list:
    """Parse Recipients sheet."""
    if "Recipients" not in wb.sheetnames:
        raise ValueError("Missing required sheet: Recipients")

    sheet = wb["Recipients"]
    headers = [cell.value for cell in sheet[1]]  # First row is header

    recipients = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:  # Skip empty rows
            continue

        recipient_dict = {}
        for i, header in enumerate(headers):
            if header is None:
                continue
            value = row[i] if i < len(row) else None

            # Convert column names to model field names
            header_clean = header.strip().lower().replace(" ", "_")

            # Handle recipient_ips (comma-separated string → list)
            if header_clean == "recipient_ips" and value:
                if isinstance(value, str):
                    recipient_dict[header_clean] = [ip.strip() for ip in value.split(",") if ip.strip()]
                else:
                    recipient_dict[header_clean] = []
            # Handle token_rotation (string → boolean)
            elif header_clean == "token_rotation":
                if isinstance(value, str):
                    recipient_dict[header_clean] = value.lower() in ("true", "yes", "1")
                else:
                    recipient_dict[header_clean] = bool(value)
            else:
                recipient_dict[header_clean] = value if value is not None else ""

        recipients.append(recipient_dict)

    return recipients


def _parse_shares_sheet(wb) -> list:
    """Parse Shares sheet (grouped by share_name)."""
    if "Shares" not in wb.sheetnames:
        raise ValueError("Missing required sheet: Shares")

    sheet = wb["Shares"]
    headers = [cell.value for cell in sheet[1]]

    # Group rows by share_name
    shares_dict = {}  # share_name → list of rows

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:  # Skip empty rows
            continue

        row_dict = {}
        for i, header in enumerate(headers):
            if header is None:
                continue
            value = row[i] if i < len(row) else None
            header_clean = header.strip().lower().replace(" ", "_")
            row_dict[header_clean] = value if value is not None else ""

        share_name = row_dict.get("share_name")
        if not share_name:
            continue

        if share_name not in shares_dict:
            shares_dict[share_name] = []
        shares_dict[share_name].append(row_dict)

    # Convert to share config format
    shares = []
    for share_name, rows in shares_dict.items():
        # Collect unique assets and recipients
        assets = []
        recipients = []

        # Get delta_share config from first row
        first_row = rows[0]
        delta_share = {
            "ext_catalog_name": first_row.get("ext_catalog_name", ""),
            "ext_schema_name": first_row.get("ext_schema_name", ""),
            "prefix_assetname": first_row.get("prefix_assetname", ""),
            "tags": [tag.strip() for tag in first_row.get("share_tags", "").split(",") if tag.strip()],
        }

        # Collect assets and recipients from all rows
        for row in rows:
            asset = row.get("asset")
            if asset and asset not in assets:
                assets.append(asset)

            recipient = row.get("recipient")
            if recipient and recipient not in recipients:
                recipients.append(recipient)

        shares.append({
            "name": share_name,
            "share_assets": assets,
            "recipients": recipients,
            "delta_share": delta_share,
            "pipelines": [],  # Will be populated from Pipelines sheet
        })

    return shares


def _parse_pipelines_sheet(wb) -> dict:
    """Parse Pipelines sheet (grouped by share_name)."""
    if "Pipelines" not in wb.sheetnames:
        logger.debug("Pipelines sheet not found - shares will have no pipelines")
        return {}

    sheet = wb["Pipelines"]
    headers = [cell.value for cell in sheet[1]]

    # Group by share_name
    pipelines_dict = {}  # share_name → list of pipeline configs

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:  # Skip empty rows
            continue

        row_dict = {}
        for i, header in enumerate(headers):
            if header is None:
                continue
            value = row[i] if i < len(row) else None
            header_clean = header.strip().lower().replace(" ", "_")
            row_dict[header_clean] = value if value is not None else ""

        share_name = row_dict.get("share_name")
        if not share_name:
            continue

        # Build pipeline config
        # Schedule is per-asset: {asset_name: {cron: ..., timezone: ...} or "continuous"}
        asset_name = row_dict.get("asset_name", "")
        schedule_type = row_dict.get("schedule_type", "CRON").upper()

        if schedule_type == "CONTINUOUS":
            schedule_value = "continuous"
        else:
            schedule_value = {
                "cron": row_dict.get("cron", "0 0 * * *"),
                "timezone": row_dict.get("timezone", "UTC"),
            }

        pipeline = {
            "name_prefix": row_dict.get("name_prefix", "pipeline"),
            "schedule": {asset_name: schedule_value},
            "notification": [n.strip() for n in row_dict.get("notification", "").split(",") if n.strip()],
            "tags": dict(item.split(":") for item in row_dict.get("tags", "").split(",") if ":" in item),
            "serverless": row_dict.get("serverless", "false").lower() in ("true", "yes", "1"),
            "scd_type": row_dict.get("scd_type", "2"),
            "key_columns": row_dict.get("key_columns", ""),
        }

        if share_name not in pipelines_dict:
            pipelines_dict[share_name] = []
        pipelines_dict[share_name].append(pipeline)

    return pipelines_dict
