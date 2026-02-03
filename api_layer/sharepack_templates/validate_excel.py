#!/usr/bin/env python3
"""
SharePack Excel File Validator

This script validates the structure and content of SharePack Excel files.
It checks for required sheets, columns, data formats, and common errors.

Usage:
    python validate_excel.py sample_sharepack.xlsx
    python validate_excel.py path/to/your_sharepack.xlsx
"""

import sys
from pathlib import Path
from typing import Dict, List, Tuple
import re

try:
    import pandas as pd
    from openpyxl import load_workbook
except ImportError:
    print("❌ Missing required packages. Install with:")
    print("   pip install pandas openpyxl")
    sys.exit(1)


class SharePackValidator:
    """Validates SharePack Excel file structure and content."""

    # Required sheets
    REQUIRED_SHEETS = ["metadata", "recipient", "share", "pipelines"]

    # Required metadata fields
    REQUIRED_METADATA_FIELDS = {
        "requestor",
        "business_line",
        "strategy",
        "delta_share_region",
        "configurator",
        "approver",
        "executive_team",
        "approver_status",
        "workspace_url",
        "servicenow",
    }

    # Required columns per sheet
    REQUIRED_COLUMNS = {
        "metadata": {"Field", "Value"},
        "recipient": {"name", "type", "recipient"},
        "share": {"name", "recipients", "share_assets", "ext_catalog_name", "ext_schema_name"},
        "pipelines": {
            "share_name",
            "name_prefix",
            "source_asset",
            "target_asset",
            "scd_type",
            "serverless",
        },
    }

    def __init__(self, file_path: str):
        """Initialize validator with Excel file path."""
        self.file_path = Path(file_path)
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.info: List[str] = []

    def validate(self) -> Tuple[bool, List[str], List[str], List[str]]:
        """
        Validate the Excel file.

        Returns:
            Tuple of (is_valid, errors, warnings, info)
        """
        # Check file exists
        if not self.file_path.exists():
            self.errors.append(f"File not found: {self.file_path}")
            return False, self.errors, self.warnings, self.info

        self.info.append(f"📁 Validating: {self.file_path.name}")
        self.info.append("")

        # Load workbook
        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception as e:
            self.errors.append(f"Failed to open Excel file: {e}")
            return False, self.errors, self.warnings, self.info

        # Validate sheets
        self._validate_sheets(sheet_names)

        # Validate each sheet
        try:
            for sheet_name in self.REQUIRED_SHEETS:
                if sheet_name in sheet_names:
                    self._validate_sheet(sheet_name)
        except Exception as e:
            self.errors.append(f"Error during validation: {e}")

        is_valid = len(self.errors) == 0
        return is_valid, self.errors, self.warnings, self.info

    def _validate_sheets(self, sheet_names: List[str]) -> None:
        """Validate required sheets exist."""
        self.info.append("📋 Sheet Validation:")

        missing_sheets = set(self.REQUIRED_SHEETS) - set(sheet_names)
        extra_sheets = set(sheet_names) - set(self.REQUIRED_SHEETS)

        if missing_sheets:
            self.errors.append(f"Missing required sheets: {', '.join(missing_sheets)}")
        else:
            self.info.append("   ✅ All required sheets present")

        if extra_sheets:
            self.warnings.append(f"Extra sheets found (will be ignored): {', '.join(extra_sheets)}")

        for sheet in self.REQUIRED_SHEETS:
            status = "✅" if sheet in sheet_names else "❌"
            self.info.append(f"   {status} {sheet}")

        self.info.append("")

    def _validate_sheet(self, sheet_name: str) -> None:
        """Validate individual sheet structure and content."""
        self.info.append(f"📄 Validating '{sheet_name}' sheet:")

        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        except Exception as e:
            self.errors.append(f"Failed to read '{sheet_name}' sheet: {e}")
            return

        # Special handling for metadata sheet (2-column format)
        if sheet_name == "metadata":
            self._validate_metadata_sheet(df)
        else:
            self._validate_data_sheet(sheet_name, df)

        self.info.append("")

    def _validate_metadata_sheet(self, df: pd.DataFrame) -> None:
        """Validate metadata sheet (Field | Value format)."""
        # Check columns
        expected_cols = {"Field", "Value"}
        actual_cols = set(df.columns)

        if not expected_cols.issubset(actual_cols):
            self.errors.append(
                f"Metadata sheet must have 'Field' and 'Value' columns. Found: {list(actual_cols)}"
            )
            return

        # Extract fields
        fields = df["Field"].dropna().str.strip().tolist()
        values = df["Value"].dropna().tolist()

        self.info.append(f"   Found {len(fields)} metadata fields")

        # Check required fields
        fields_set = set(fields)
        missing_fields = self.REQUIRED_METADATA_FIELDS - fields_set

        if missing_fields:
            self.errors.append(f"Missing required metadata fields: {', '.join(sorted(missing_fields))}")
        else:
            self.info.append("   ✅ All required metadata fields present")

        # Validate field values
        field_values = dict(zip(fields, values))
        self._validate_metadata_values(field_values)

    def _validate_metadata_values(self, field_values: Dict[str, any]) -> None:
        """Validate metadata field values."""
        # Strategy
        if "strategy" in field_values:
            strategy = str(field_values["strategy"]).upper()
            if strategy not in ["NEW", "UPDATE"]:
                self.errors.append(f"Invalid strategy: '{strategy}'. Must be 'NEW' or 'UPDATE'")
            else:
                self.info.append(f"   ✅ Strategy: {strategy}")

        # Delta share region
        if "delta_share_region" in field_values:
            region = str(field_values["delta_share_region"]).upper()
            if region not in ["AM", "EMEA"]:
                self.errors.append(f"Invalid delta_share_region: '{region}'. Must be 'AM' or 'EMEA'")
            else:
                self.info.append(f"   ✅ Region: {region}")

        # Approver status
        if "approver_status" in field_values:
            status = str(field_values["approver_status"]).lower()
            valid_statuses = ["approved", "declined", "request_more_info", "pending"]
            if status not in valid_statuses:
                self.errors.append(
                    f"Invalid approver_status: '{status}'. Must be one of: {', '.join(valid_statuses)}"
                )

        # Email validation
        email_fields = ["requestor", "configurator", "approver", "executive_team", "contact_email"]
        for field in email_fields:
            if field in field_values:
                value = str(field_values[field])
                if "@" in value and not self._is_valid_email(value):
                    self.warnings.append(f"Potentially invalid email in {field}: {value}")

        # Workspace URL
        if "workspace_url" in field_values:
            url = str(field_values["workspace_url"])
            if not url.startswith("https://"):
                self.errors.append(f"workspace_url must start with https://. Found: {url}")
            elif "azuredatabricks.net" not in url and "cloud.databricks.com" not in url:
                self.warnings.append(f"Workspace URL doesn't match expected patterns: {url}")
            else:
                self.info.append(f"   ✅ Workspace URL: {url}")

        # ServiceNow
        if "servicenow" in field_values:
            sn = str(field_values["servicenow"])
            if not sn or sn.lower() in ["nan", "none", ""]:
                self.errors.append("servicenow field is required")
            else:
                self.info.append(f"   ✅ ServiceNow: {sn}")

    def _validate_data_sheet(self, sheet_name: str, df: pd.DataFrame) -> None:
        """Validate data sheet (recipient, share, pipelines)."""
        # Check required columns
        required_cols = self.REQUIRED_COLUMNS.get(sheet_name, set())
        actual_cols = set(df.columns)

        missing_cols = required_cols - actual_cols
        if missing_cols:
            self.errors.append(f"Sheet '{sheet_name}' missing required columns: {', '.join(sorted(missing_cols))}")
        else:
            self.info.append(f"   ✅ All required columns present ({len(required_cols)} required)")

        # Check for data
        row_count = len(df)
        if row_count == 0:
            self.warnings.append(f"Sheet '{sheet_name}' has no data rows")
        else:
            self.info.append(f"   ✅ Found {row_count} data row(s)")

        # Sheet-specific validation
        if sheet_name == "recipient" and row_count > 0:
            self._validate_recipients(df)
        elif sheet_name == "share" and row_count > 0:
            self._validate_shares(df)
        elif sheet_name == "pipelines" and row_count > 0:
            self._validate_pipelines(df)

    def _validate_recipients(self, df: pd.DataFrame) -> None:
        """Validate recipient data."""
        for idx, row in df.iterrows():
            # Type validation
            if "type" in row:
                rec_type = str(row["type"]).upper()
                if rec_type not in ["D2O", "D2D"]:
                    self.errors.append(f"Row {idx + 2}: Invalid recipient type '{rec_type}'. Must be D2O or D2D")

                # D2D requires recipient_databricks_org
                if rec_type == "D2D" and pd.isna(row.get("recipient_databricks_org")):
                    self.errors.append(f"Row {idx + 2}: D2D recipient requires recipient_databricks_org")

            # Email validation
            if "recipient" in row and not pd.isna(row["recipient"]):
                email = str(row["recipient"])
                if "@" in email and not self._is_valid_email(email):
                    self.warnings.append(f"Row {idx + 2}: Potentially invalid email: {email}")

    def _validate_shares(self, df: pd.DataFrame) -> None:
        """Validate share data."""
        for idx, row in df.iterrows():
            # Share assets should be comma-separated
            if "share_assets" in row and not pd.isna(row["share_assets"]):
                assets = str(row["share_assets"])
                asset_list = [a.strip() for a in assets.split(",")]

                # Check 3-part names (catalog.schema.table)
                for asset in asset_list:
                    parts = asset.split(".")
                    if len(parts) != 3:
                        self.warnings.append(
                            f"Row {idx + 2}: Share asset should be 3-part name (catalog.schema.table): {asset}"
                        )

    def _validate_pipelines(self, df: pd.DataFrame) -> None:
        """Validate pipeline data."""
        for idx, row in df.iterrows():
            # Source asset validation (3-part name)
            if "source_asset" in row and not pd.isna(row["source_asset"]):
                source = str(row["source_asset"])
                parts = source.split(".")
                if len(parts) != 3:
                    self.errors.append(
                        f"Row {idx + 2}: source_asset must be 3-part name (catalog.schema.table): {source}"
                    )

            # SCD type validation
            if "scd_type" in row and not pd.isna(row["scd_type"]):
                scd_type = str(row["scd_type"])
                if scd_type not in ["1", "2"]:
                    self.errors.append(f"Row {idx + 2}: scd_type must be '1' or '2', found: {scd_type}")

            # Serverless validation
            if "serverless" in row and not pd.isna(row["serverless"]):
                serverless = str(row["serverless"]).lower()
                if serverless not in ["true", "false"]:
                    self.warnings.append(f"Row {idx + 2}: serverless should be true/false, found: {serverless}")

            # Schedule cron validation (if present)
            if "schedule_cron" in row and not pd.isna(row["schedule_cron"]):
                cron = str(row["schedule_cron"])
                if cron.lower() != "continuous":
                    # Basic cron format check (6 fields)
                    cron_parts = cron.split()
                    if len(cron_parts) != 6:
                        self.warnings.append(
                            f"Row {idx + 2}: Cron expression should have 6 fields (Quartz format): {cron}"
                        )

            # Tags format validation (semicolon-separated key:value)
            if "tags" in row and not pd.isna(row["tags"]):
                tags = str(row["tags"])
                if ";" in tags:
                    tag_list = [t.strip() for t in tags.split(";")]
                    for tag in tag_list:
                        if ":" not in tag:
                            self.warnings.append(
                                f"Row {idx + 2}: Tag should be key:value format: {tag}"
                            )

    def _is_valid_email(self, email: str) -> bool:
        """Basic email validation."""
        pattern = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
        return re.match(pattern, email) is not None


def print_results(is_valid: bool, errors: List[str], warnings: List[str], info: List[str]) -> None:
    """Print validation results."""
    # Print info
    for line in info:
        print(line)

    # Print warnings
    if warnings:
        print("⚠️  WARNINGS:")
        for warning in warnings:
            print(f"   {warning}")
        print()

    # Print errors
    if errors:
        print("❌ ERRORS:")
        for error in errors:
            print(f"   {error}")
        print()

    # Summary
    print("=" * 70)
    if is_valid:
        if warnings:
            print("✅ VALIDATION PASSED (with warnings)")
        else:
            print("✅ VALIDATION PASSED")
        print("\n📤 Excel file is ready to upload!")
    else:
        print("❌ VALIDATION FAILED")
        print(f"\n🔧 Fix {len(errors)} error(s) before uploading")
    print("=" * 70)


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("Usage: python validate_excel.py <excel_file.xlsx>")
        print("\nExample:")
        print("  python validate_excel.py sample_sharepack.xlsx")
        sys.exit(1)

    file_path = sys.argv[1]

    # Validate
    validator = SharePackValidator(file_path)
    is_valid, errors, warnings, info = validator.validate()

    # Print results
    print_results(is_valid, errors, warnings, info)

    # Exit code
    sys.exit(0 if is_valid else 1)


if __name__ == "__main__":
    main()
